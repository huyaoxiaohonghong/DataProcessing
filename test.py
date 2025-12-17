import pandas as pd
import os

# ================= 配置区域 =================
folder_path = r'text'  # 文件夹路径
source_file = '专用设备.xlsx'  # 源文件
#source_file = '通用设备.xlsx'  # 源文件
#source_file = '家具和用具.xlsx'  # 源文件
#source_file = '房屋和构筑物.xlsx'  # 源文件
target_file = '固定资产卡片.xlsx'  # 目标文件
lookup_file = '对照表.xlsx'  # 对照表文件
target_sheet = '主表和子表导入'  # 目标Sheet
# ===========================================

# 字段映射关系（源文件字段 -> 目标文件字段）
# 直接映射（不需要对照表转换）
direct_mapping = {
    '资产名称': '资产名称',
    '取得日期': '开始使用日期',
    '折旧/摊销年限(月)': '使用月限',
    '资产原值(元)': '原值',
    '累计折旧/摊销(元)': '累计折旧',
    '数量/面积': '数量',
    '记账日期': '入账日期',
    '已提折旧/摊销月数': '已计提月份',
    '净值(元)': '净值',
}

# 需要通过对照表转换的映射（源字段 -> (对照表Sheet, 对照表名称列, 对照表编码列, 目标字段)）
lookup_mapping = {
    '取得方式': ('增加方式', '增加方式名称', '增加方式编码', '增加方式编码'),
    '资产分类': ('资产类别', '资产类别名称', '资产类别编码', '资产类别编码'),
    '资产用途': ('资产用途', '资产用途名称', '资产用途编码', '资产用途编码'),
    '资产状态': ('使用状态', '使用状态名称', '使用状态编码', '使用状态编码'),
    '管理部门': ('科室', '管理部门', '管理科室编码', '管理科室编码'),
    '资产门类': ('资产门类', '资产门类名称', '资产门类编码', None),  # 用于生成主表主键
}

print("=" * 60)
print("开始导入数据：房屋和构筑物 -> 固定资产卡片")
print("=" * 60)

try:
    # 1. 读取对照表
    lookup_path = os.path.join(folder_path, lookup_file)
    print(f"\n正在读取对照表: {lookup_path}")
    lookup_xls = pd.ExcelFile(lookup_path)
    
    # 构建对照字典
    lookup_dicts = {}
    for source_field, (sheet_name, name_col, code_col, target_field) in lookup_mapping.items():
        # 读取时指定dtype为str，保留前导0
        df = pd.read_excel(lookup_xls, sheet_name=sheet_name, dtype={code_col: str})
        # 创建名称到编码的映射字典，保留编码的原始格式
        lookup_dicts[source_field] = dict(zip(
            df[name_col].astype(str).str.strip(), 
            df[code_col].astype(str).str.strip()
        ))
        print(f"  - 加载 '{sheet_name}' 对照表: {len(lookup_dicts[source_field])} 条记录")
    
    # 2. 读取源文件
    source_path = os.path.join(folder_path, source_file)
    print(f"\n正在读取源文件: {source_path}")
    source_df = pd.read_excel(source_path, sheet_name='Sheet1')
    print(f"源文件读取成功，共 {len(source_df)} 行数据")

    # 3. 读取目标文件
    target_path = os.path.join(folder_path, target_file)
    print(f"\n正在读取目标文件: {target_path}")
    target_xls = pd.ExcelFile(target_path)
    
    # 读取目标文件的所有Sheet
    all_sheets = {}
    for sheet_name in target_xls.sheet_names:
        all_sheets[sheet_name] = pd.read_excel(target_xls, sheet_name=sheet_name)
    
    # 获取目标Sheet
    if target_sheet in all_sheets:
        target_df = all_sheets[target_sheet]
        print(f"目标Sheet '{target_sheet}' 当前有 {len(target_df)} 行数据")
        print(f"目标Sheet 字段数: {len(target_df.columns)}")
    else:
        target_df = pd.DataFrame()
        print(f"目标Sheet '{target_sheet}' 不存在，将创建新Sheet")

    # 获取目标表的所有列名
    if len(target_df.columns) == 0:
        print("\n✗ 错误：目标Sheet没有字段结构")
        print("请确保目标文件中至少有表头")
        exit()
    
    target_columns = list(target_df.columns)
    print(f"目标表字段: {target_columns[:5]}... (共{len(target_columns)}个)")
    
    # 4. 创建新的数据行
    new_rows = []
    mapped_fields = []
    lookup_stats = {field: {'success': 0, 'fail': 0} for field in lookup_mapping.keys()}
    primary_key_counter = {}  # 用于记录每个资产门类编码的计数器
    
    for idx, row in source_df.iterrows():
        # 创建空行，包含目标表的所有列
        new_row = {col: None for col in target_columns}
        
        # 先获取资产门类编码（用于生成主表主键）
        asset_category_code = None
        if '资产门类' in source_df.columns:
            source_value = str(row['资产门类']).strip() if pd.notna(row['资产门类']) else None
            if source_value:
                asset_category_code = lookup_dicts['资产门类'].get(source_value)
                if asset_category_code:
                    lookup_stats['资产门类']['success'] += 1
                else:
                    lookup_stats['资产门类']['fail'] += 1
        
        # 生成主表主键（资产门类编码 + 序号）
        if asset_category_code and '主表主键' in target_columns:
            # 初始化该资产门类的计数器
            if asset_category_code not in primary_key_counter:
                primary_key_counter[asset_category_code] = 1
            else:
                primary_key_counter[asset_category_code] += 1
            
            # 生成主表主键：资产门类编码 + 4位序号（如：010001, 010002...）
            sequence = str(primary_key_counter[asset_category_code]).zfill(4)
            primary_key = f"{asset_category_code}{sequence}"
            new_row['主表主键'] = primary_key
            
            if '主表主键' not in mapped_fields:
                mapped_fields.append('主表主键')
        
        # 处理直接映射字段
        for source_col, target_col in direct_mapping.items():
            if source_col in source_df.columns and target_col in target_columns:
                if pd.notna(row[source_col]):
                    new_row[target_col] = row[source_col]
                    if target_col not in mapped_fields:
                        mapped_fields.append(target_col)
        
        # 处理需要对照表转换的字段（排除资产门类，已在上面处理）
        for source_col, (sheet_name, name_col, code_col, target_col) in lookup_mapping.items():
            if source_col == '资产门类':  # 资产门类已处理，跳过
                continue
            if source_col in source_df.columns and target_col in target_columns:
                source_value = str(row[source_col]).strip() if pd.notna(row[source_col]) else None
                if source_value:
                    # 在对照表中查找对应的编码
                    code = lookup_dicts[source_col].get(source_value)
                    if code:
                        new_row[target_col] = code
                        lookup_stats[source_col]['success'] += 1
                        if target_col not in mapped_fields:
                            mapped_fields.append(target_col)
                    else:
                        lookup_stats[source_col]['fail'] += 1
        
        # 处理"多资金来源"字段
        if '多资金来源' in target_columns:
            # 检查"非财政拨款(元)"和"财政拨款(元)"是否都不为空且大于0
            non_fiscal = 0
            fiscal = 0
            
            if '非财政拨款(元)' in source_df.columns and pd.notna(row['非财政拨款(元)']):
                try:
                    non_fiscal = float(row['非财政拨款(元)'])
                except (ValueError, TypeError):
                    non_fiscal = 0
            
            if '财政拨款(元)' in source_df.columns and pd.notna(row['财政拨款(元)']):
                try:
                    fiscal = float(row['财政拨款(元)'])
                except (ValueError, TypeError):
                    fiscal = 0
            
            # 如果两个字段都大于0，则多资金来源为1，否则为0
            if non_fiscal > 0 and fiscal > 0:
                new_row['多资金来源'] = 1
            else:
                new_row['多资金来源'] = 0
            
            if '多资金来源' not in mapped_fields:
                mapped_fields.append('多资金来源')
        
        # 设置默认值
        if '原始卡片' in target_columns:
            new_row['原始卡片'] = 1
            if '原始卡片' not in mapped_fields:
                mapped_fields.append('原始卡片')
        
        if '多使用科室' in target_columns:
            new_row['多使用科室'] = 0
            if '多使用科室' not in mapped_fields:
                mapped_fields.append('多使用科室')
        
        if '账簿编码' in target_columns:
            new_row['账簿编码'] = 'BC20251125155451'
            if '账簿编码' not in mapped_fields:
                mapped_fields.append('账簿编码')
        
        if '会计期间' in target_columns:
            new_row['会计期间'] = '2025-11'
            if '会计期间' not in mapped_fields:
                mapped_fields.append('会计期间')
        
        if '会计期间text' in target_columns:
            new_row['会计期间text'] = '2025-11'
            if '会计期间text' not in mapped_fields:
                mapped_fields.append('会计期间text')
        
        if '净残值率' in target_columns:
            new_row['净残值率'] = '0%'
            if '净残值率' not in mapped_fields:
                mapped_fields.append('净残值率')
        
        if '净残值' in target_columns:
            new_row['净残值'] = 0
            if '净残值' not in mapped_fields:
                mapped_fields.append('净残值')
        
        if '折旧方法编码' in target_columns:
            new_row['折旧方法编码'] = 'NianXianPingJunFa'
            if '折旧方法编码' not in mapped_fields:
                mapped_fields.append('折旧方法编码')
        
        # 计算字段
        if '卡片编号' in target_columns and new_row.get('主表主键'):
            new_row['卡片编号'] = new_row['主表主键']
            if '卡片编号' not in mapped_fields:
                mapped_fields.append('卡片编号')
        
        if '资产编码' in target_columns and new_row.get('主表主键'):
            new_row['资产编码'] = new_row['主表主键']
            if '资产编码' not in mapped_fields:
                mapped_fields.append('资产编码')
        
        if '单价' in target_columns and new_row.get('原值') is not None:
            new_row['单价'] = new_row['原值']
            if '单价' not in mapped_fields:
                mapped_fields.append('单价')
        
        if '净额' in target_columns and new_row.get('净值') is not None:
            new_row['净额'] = new_row['净值']
            if '净额' not in mapped_fields:
                mapped_fields.append('净额')
        
        if '使用年限' in target_columns and new_row.get('使用月限') is not None:
            try:
                months = float(new_row['使用月限'])
                new_row['使用年限'] = months / 12
                if '使用年限' not in mapped_fields:
                    mapped_fields.append('使用年限')
            except (ValueError, TypeError):
                pass
        
        # 处理空值默认值
        if '已计提月份' in target_columns:
            if new_row.get('已计提月份') is None or pd.isna(new_row.get('已计提月份')):
                new_row['已计提月份'] = 0
        
        if '累计折旧' in target_columns:
            if new_row.get('累计折旧') is None or pd.isna(new_row.get('累计折旧')):
                new_row['累计折旧'] = 0
        
        new_rows.append(new_row)

    # 将新数据转换为DataFrame
    new_df = pd.DataFrame(new_rows, columns=target_columns)
    
    print(f"\n成功映射的字段 ({len(mapped_fields)} 个):")
    print("  直接映射:")
    for field in direct_mapping.values():
        if field in mapped_fields:
            source_field = [k for k, v in direct_mapping.items() if v == field][0]
            print(f"    - {source_field} → {field}")
    
    print("  对照表转换:")
    for source_field, (sheet_name, name_col, code_col, target_field) in lookup_mapping.items():
        success = lookup_stats[source_field]['success']
        fail = lookup_stats[source_field]['fail']
        if source_field == '资产门类':
            print(f"    - {source_field} → 主表主键 (成功:{success}, 失败:{fail})")
        elif target_field and target_field in mapped_fields:
            print(f"    - {source_field} → {target_field} (成功:{success}, 失败:{fail})")
    
    print(f"\n准备导入 {len(new_df)} 行数据")
    print(f"新数据包含 {len(new_df.columns)} 个字段（与目标表一致）")

    # 5. 合并数据到主表和子表导入
    result_df = pd.concat([target_df, new_df], ignore_index=True)
    all_sheets[target_sheet] = result_df
    
    # 6. 生成"使用科室导入"Sheet的数据
    print("\n正在生成使用科室导入数据...")
    
    # 读取科室对照表，建立编码到名称的映射
    dept_df = pd.read_excel(lookup_xls, sheet_name='科室', dtype={'管理科室编码': str})
    dept_code_to_name = dict(zip(
        dept_df['管理科室编码'].astype(str).str.strip(),
        dept_df['管理科室名称'].astype(str).str.strip()
    ))
    
    # 获取"使用科室导入"Sheet的原有数据和列结构
    dept_import_sheet = '使用科室导入'
    if dept_import_sheet in all_sheets:
        dept_import_df = all_sheets[dept_import_sheet]
        dept_columns = list(dept_import_df.columns)
    else:
        # 如果不存在，创建默认列结构
        dept_columns = ['主表主键', '会计期间', '编码对照id', '使用科室编码', '使用科室名称', '折旧承担比率']
        dept_import_df = pd.DataFrame(columns=dept_columns)
    
    # 为每条主表记录生成对应的使用科室导入记录
    dept_import_rows = []
    for _, row in new_df.iterrows():
        primary_key = row.get('主表主键')
        dept_code = row.get('管理科室编码')
        
        if primary_key and dept_code:
            # 查找科室名称
            dept_name = dept_code_to_name.get(str(dept_code).strip(), '')
            
            dept_row = {col: None for col in dept_columns}
            dept_row['主表主键'] = primary_key
            dept_row['会计期间'] = '2025-11'
            dept_row['编码对照id'] = dept_code  # 编码对照id等于使用科室编码
            dept_row['使用科室编码'] = dept_code
            dept_row['使用科室名称'] = dept_name
            dept_row['折旧承担比率'] = '100%'  # 默认值为100%
            
            dept_import_rows.append(dept_row)
    
    # 创建新的使用科室导入DataFrame
    new_dept_df = pd.DataFrame(dept_import_rows, columns=dept_columns)
    
    # 合并到原有数据
    result_dept_df = pd.concat([dept_import_df, new_dept_df], ignore_index=True)
    all_sheets[dept_import_sheet] = result_dept_df
    
    print(f"  - 生成了 {len(new_dept_df)} 条使用科室导入记录")
    
    # 7. 生成"资金来源导入"Sheet的数据
    print("\n正在生成资金来源导入数据...")
    
    # 读取项目对照表
    # 流程：源文件项目代码 → 对照表项目编码 → 对照表项目名称
    project_df = pd.read_excel(lookup_xls, sheet_name='项目', dtype={'项目编码': str})
    
    # 建立映射字典
    # 如果对照表有"项目代码"列，则：项目代码 -> 项目编码
    # 如果没有，则假设源文件的项目代码就是对照表的项目编码
    if '项目代码' in project_df.columns:
        # 项目代码 -> 项目编码的映射
        project_code_to_code = dict(zip(
            project_df['项目代码'].astype(str).str.strip(),
            project_df['项目编码'].astype(str).str.strip()
        ))
    else:
        # 如果没有项目代码列，假设源文件的项目代码就是项目编码
        project_code_to_code = None
    
    # 项目编码 -> 项目名称的映射
    project_code_to_name = dict(zip(
        project_df['项目编码'].astype(str).str.strip(),
        project_df['项目名称'].astype(str).str.strip()
    ))
    
    # 获取"资金来源导入"Sheet的原有数据和列结构
    fund_import_sheet = '资金来源导入'
    if fund_import_sheet in all_sheets:
        fund_import_df = all_sheets[fund_import_sheet]
        fund_columns = list(fund_import_df.columns)
    else:
        # 如果不存在，创建默认列结构
        fund_columns = ['主表主键', '会计期间', '资金来源编码', '资金来源名称', '原值', '累计折旧', '净值', '项目编码', '项目名称']
        fund_import_df = pd.DataFrame(columns=fund_columns)
    
    # 为每条主表记录生成对应的资金来源导入记录
    fund_import_rows = []
    for idx, main_row in new_df.iterrows():
        primary_key = main_row.get('主表主键')
        if not primary_key:
            continue
        
        # 获取源文件对应行的数据
        source_row = source_df.iloc[idx]
        
        # 流程1：源文件项目代码 → 对照表项目代码 → 对照表项目编码 → 目标文件项目编码
        # 流程2：目标文件项目编码 → 对照表项目编码 → 对照表项目名称 → 目标文件项目名称
        project_code = None
        project_name = None
        
        if '项目代码' in source_df.columns and pd.notna(source_row['项目代码']):
            source_project_code = str(source_row['项目代码']).strip()
            
            # 步骤1：源文件项目代码 → 对照表项目编码
            if project_code_to_code:
                # 如果有项目代码到项目编码的映射
                project_code = project_code_to_code.get(source_project_code, None)
            else:
                # 如果没有映射，假设源文件的项目代码就是项目编码
                project_code = source_project_code
            
            # 步骤2：对照表项目编码 → 对照表项目名称
            if project_code:
                project_name = project_code_to_name.get(project_code, None)
        
        # 获取财政拨款和非财政拨款的值
        fiscal = 0
        non_fiscal = 0
        
        if '财政拨款(元)' in source_df.columns and pd.notna(source_row['财政拨款(元)']):
            try:
                fiscal = float(source_row['财政拨款(元)'])
            except (ValueError, TypeError):
                fiscal = 0
        
        if '非财政拨款(元)' in source_df.columns and pd.notna(source_row['非财政拨款(元)']):
            try:
                non_fiscal = float(source_row['非财政拨款(元)'])
            except (ValueError, TypeError):
                non_fiscal = 0
        
        # 获取累计折旧和净值
        accumulated_depreciation = source_row.get('累计折旧/摊销(元)', 0) if '累计折旧/摊销(元)' in source_df.columns else 0
        net_value = source_row.get('净值(元)', 0) if '净值(元)' in source_df.columns else 0
        
        # 情况1：只有非财政拨款
        if (fiscal == 0 or pd.isna(fiscal)) and non_fiscal > 0:
            fund_row = {col: None for col in fund_columns}
            fund_row['主表主键'] = primary_key
            fund_row['会计期间'] = '2025-11'
            fund_row['资金来源编码'] = '0005'
            fund_row['资金来源名称'] = '非财政拨款'
            fund_row['原值'] = non_fiscal
            fund_row['累计折旧'] = accumulated_depreciation
            fund_row['净值'] = net_value
            fund_row['项目编码'] = project_code
            fund_row['项目名称'] = project_name
            fund_import_rows.append(fund_row)
        
        # 情况2：只有财政拨款
        elif (non_fiscal == 0 or pd.isna(non_fiscal)) and fiscal > 0:
            fund_row = {col: None for col in fund_columns}
            fund_row['主表主键'] = primary_key
            fund_row['会计期间'] = '2025-11'
            fund_row['资金来源编码'] = '0002'
            fund_row['资金来源名称'] = '财政项目拨款资金'
            fund_row['原值'] = fiscal
            fund_row['累计折旧'] = accumulated_depreciation
            fund_row['净值'] = net_value
            fund_row['项目编码'] = project_code
            fund_row['项目名称'] = project_name
            fund_import_rows.append(fund_row)
        
        # 情况3：两者都有
        elif non_fiscal > 0 and fiscal > 0:
            # 第一条：非财政拨款
            fund_row1 = {col: None for col in fund_columns}
            fund_row1['主表主键'] = primary_key
            fund_row1['会计期间'] = '2025-11'
            fund_row1['资金来源编码'] = '0005'
            fund_row1['资金来源名称'] = '非财政拨款'
            fund_row1['原值'] = non_fiscal + fiscal  # 原值为两者之和
            fund_row1['累计折旧'] = 0  # 非财政拨款的累计折旧为0
            fund_row1['净值'] = 0  # 非财政拨款的净值为0
            fund_row1['项目编码'] = project_code
            fund_row1['项目名称'] = project_name
            fund_import_rows.append(fund_row1)
            
            # 第二条：财政项目拨款资金
            fund_row2 = {col: None for col in fund_columns}
            fund_row2['主表主键'] = primary_key
            fund_row2['会计期间'] = '2025-11'
            fund_row2['资金来源编码'] = '0002'
            fund_row2['资金来源名称'] = '财政项目拨款资金'
            fund_row2['原值'] = non_fiscal + fiscal  # 原值为两者之和
            fund_row2['累计折旧'] = accumulated_depreciation  # 财政拨款的累计折旧等于源文件值
            fund_row2['净值'] = net_value  # 财政拨款的净值等于源文件值
            fund_row2['项目编码'] = project_code
            fund_row2['项目名称'] = project_name
            fund_import_rows.append(fund_row2)
    
    # 创建新的资金来源导入DataFrame
    new_fund_df = pd.DataFrame(fund_import_rows, columns=fund_columns)
    
    # 处理累计折旧字段的空值默认值
    if '累计折旧' in new_fund_df.columns:
        new_fund_df['累计折旧'] = new_fund_df['累计折旧'].fillna(0)
    
    # 合并到原有数据
    result_fund_df = pd.concat([fund_import_df, new_fund_df], ignore_index=True)
    all_sheets[fund_import_sheet] = result_fund_df
    
    print(f"  - 生成了 {len(new_fund_df)} 条资金来源导入记录")

    # 8. 保存到新文件
    output_filename = '固定资产卡片_已导入.xlsx'
    output_path = os.path.join(folder_path, output_filename)
    print(f"\n正在保存到: {output_path}")
    
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import numbers
        
        # 先保存Excel文件（保留空值，不做替换）
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in all_sheets.items():
                # 不转换为字符串，保留原始数据类型和空值
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 重新打开文件，设置单元格格式
        print("正在设置单元格格式...")
        wb = load_workbook(output_path)
        
        # 需要设置为日期格式的字段
        date_fields = ['会计期间', '会计期间text', '开始使用日期', '入账日期']
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 获取表头，找到日期字段的列索引
            header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            date_columns = []
            for idx, header in enumerate(header_row, start=1):
                if header in date_fields:
                    date_columns.append(idx)
            
            # 遍历所有单元格，设置格式
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if col_idx in date_columns and row_idx > 1:  # 跳过表头
                        # 日期字段设置为日期格式
                        cell.number_format = 'YYYY-MM-DD'
                    else:
                        # 其他字段设置为文本格式
                        cell.number_format = numbers.FORMAT_TEXT
        
        wb.save(output_path)
        wb.close()
        
        print(f"\n✓ 数据导入成功！")
        print(f"  - 导入了 {len(new_df)} 行数据")
        print(f"  - 目标Sheet '{target_sheet}' 现在共有 {len(result_df)} 行数据")
        print(f"  - 所有单元格已设置为文本格式")
        print(f"  - 结果已保存到: {output_filename}")
        print("=" * 60)
    except PermissionError:
        print(f"\n✗ 无法保存文件，请确保以下文件未被打开：")
        print(f"  - {output_filename}")
        print(f"  - {target_file}")
        print("\n请关闭这些文件后重新运行脚本。")

except FileNotFoundError as e:
    print(f"\n✗ 错误：找不到文件")
    print(f"  {e}")
except Exception as e:
    print(f"\n✗ 处理过程中出错: {e}")
    import traceback
    traceback.print_exc()

print("\n处理完成！")
