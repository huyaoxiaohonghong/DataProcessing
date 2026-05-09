"""
数据处理服务
Excel parsing and data processing service
"""
import io
import logging
import re
import time
from itertools import product

from django.core.files.base import ContentFile
from django.utils import timezone
from openpyxl import load_workbook, Workbook

from .models import (
    DataMapping, MappingField, MappingTargetSheet,
    SheetLineage, ProcessingTask, TaskSheetResult,
)
from .services_utils import append_task_error
from .services_exporter import (
    build_workbook_fresh, build_workbook_from_template, sanitize_filename,
)
from utils.safe_eval import safe_eval_expr

logger = logging.getLogger('apps')

# 预编译正则
_FIELD_PATTERN = re.compile(r'\{([^}]+)\}')
# 跨 sheet 引用语法：{sheet_name.字段}  注意：与普通 {字段} 共用同一模式，区分通过是否包含 '.'


class ExcelService:
    """Excel 文件处理服务"""

    @staticmethod
    def _read_file_to_bytes(file_obj):
        """从存储后端读取文件到 BytesIO（兼容本地和 S3 存储）"""
        file_field = file_obj.file
        file_field.open('rb')
        try:
            content = file_field.read()
        finally:
            file_field.close()
        return io.BytesIO(content)

    @staticmethod
    def parse_file_fields(file_obj):
        """
        解析 Excel 文件，获取所有 Sheet 及其字段
        返回格式: [{'sheet_name': 'Sheet1', 'fields': [{'name': 'col1', 'index': 0}, ...]}]
        """
        file_bytes = ExcelService._read_file_to_bytes(file_obj)
        wb = load_workbook(file_bytes, read_only=True, data_only=True)

        try:
            result = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                fields = []

                first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if first_row:
                    for index, cell_value in enumerate(first_row):
                        if cell_value is not None:
                            fields.append({'name': str(cell_value), 'index': index})

                result.append({'sheet_name': sheet_name, 'fields': fields})

            return result
        finally:
            wb.close()

    @staticmethod
    def get_sheet_data(file_obj, sheet_name=None):
        """获取 Sheet 的所有数据"""
        file_bytes = ExcelService._read_file_to_bytes(file_obj)
        wb = load_workbook(file_bytes, read_only=True, data_only=True)

        try:
            ws = wb[sheet_name] if sheet_name else wb.active
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                return {'headers': [], 'rows': []}

            headers = [str(cell) if cell else '' for cell in rows[0]]
            data_rows = [list(row) for row in rows[1:]]
            return {'headers': headers, 'rows': data_rows}
        finally:
            wb.close()

    @staticmethod
    def read_all_sheets(file_obj):
        """读取文件所有 Sheet 数据，返回 {sheet_name: {headers, rows}}"""
        file_bytes = ExcelService._read_file_to_bytes(file_obj)
        wb = load_workbook(file_bytes, read_only=True, data_only=True)

        try:
            sheets = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    headers = [str(cell) if cell else '' for cell in rows[0]]
                    data_rows = [list(row) for row in rows[1:]]
                    sheets[sheet_name] = {'headers': headers, 'rows': data_rows}
            return sheets
        finally:
            wb.close()


# ---------------------------------------------------------------------------
# 血缘拓扑排序
# ---------------------------------------------------------------------------

def topo_sort_sheets(sheets, edges):
    """根据血缘边对目标 sheet 做拓扑排序。

    Contract (see Requirements 5.1, 5.4, 8.1):

    - **幂等** (Req 5.4): 相同 ``(sheets, edges)`` 两次调用返回相同序列。入度为 0 的节点
      按 ``(sort_order, id)`` 升序稳定出队（最小堆），确保中间入队的节点也参与全局排序。
    - **长度不变** (Req 5.1): ``len(output) == len(sheets)`` 恒成立。
    - **拓扑有效** (Req 8.1): 无环时 ∀ (u, d) ∈ edges 且 u, d ∈ sheets.id 时，
      ``index_of(u) < index_of(d)``。
    - **环的退化**: 存在环时返回按 ``(sort_order, id)`` 升序的兜底序列并记录 ``logger.warning``。
    - **边过滤**: ``u == d``（自环）、``u ∉ sheets``、``d ∉ sheets`` 的边被静默忽略，不计入度。

    Args:
        sheets: List[MappingTargetSheet] 目标 sheet 集合
        edges: List[(upstream_id, downstream_id)] 血缘边

    Returns:
        List[MappingTargetSheet] 按拓扑顺序（上游在前）或按 (sort_order, id) 升序的兜底序列。
    """
    import heapq
    from collections import defaultdict

    id_map = {s.id: s for s in sheets}
    indeg = defaultdict(int)
    graph = defaultdict(list)
    for up, down in edges:
        if up not in id_map or down not in id_map or up == down:
            continue
        graph[up].append(down)
        indeg[down] += 1

    # 最小堆：按 (sort_order, id) 全局稳定出队（即使中间才入队的节点也能稳定）
    heap = [
        (id_map[s.id].sort_order, s.id)
        for s in sheets if indeg[s.id] == 0
    ]
    heapq.heapify(heap)

    ordered_ids = []
    while heap:
        _, nid = heapq.heappop(heap)
        ordered_ids.append(nid)
        for nxt in graph[nid]:
            indeg[nxt] -= 1
            if indeg[nxt] == 0:
                heapq.heappush(heap, (id_map[nxt].sort_order, nxt))

    if len(ordered_ids) != len(sheets):
        logger.warning('Sheet 血缘中检测到环，已退化为 sort_order 顺序')
        return sorted(sheets, key=lambda s: (s.sort_order, s.id))

    return [id_map[i] for i in ordered_ids]


def detect_cycle(sheets, edges) -> bool:
    """检测 ``(sheets, edges)`` 构成的有向图是否包含环。

    与 ``topo_sort_sheets`` 语义等价：有环 ⇔ 拓扑序列长度 < |sheets|。
    独立暴露该函数，供 ``Data_Processing_Service`` 决定是否向 ``task.error_message``
    追加 "检测到 Sheet 血缘环" 警告（Req 5.2），以及供 ``Lineage_Query_API`` 复用。

    Args:
        sheets: List[MappingTargetSheet]
        edges: List[(upstream_id, downstream_id)]

    Returns:
        ``True`` 表示存在环；``False`` 表示无环（含 DAG + 被完全过滤掉的边集合）。
    """
    from collections import defaultdict, deque

    id_map = {s.id: s for s in sheets}
    indeg = defaultdict(int)
    graph = defaultdict(list)
    for up, down in edges:
        if up not in id_map or down not in id_map or up == down:
            continue
        graph[up].append(down)
        indeg[down] += 1

    queue = deque(s.id for s in sheets if indeg[s.id] == 0)
    visited_count = 0
    while queue:
        nid = queue.popleft()
        visited_count += 1
        for nxt in graph[nid]:
            indeg[nxt] -= 1
            if indeg[nxt] == 0:
                queue.append(nxt)

    return visited_count != len(sheets)


# ---------------------------------------------------------------------------
# 数据处理服务
# ---------------------------------------------------------------------------

class DataProcessingService:
    """数据处理服务"""

    PROGRESS_SAVE_INTERVAL = 100

    AGG_FUNCS = {
        'sum': lambda xs: sum(_to_number(x) or 0 for x in xs),
        'count': lambda xs: len([x for x in xs if x is not None]),
        'avg': lambda xs: (sum(_to_number(x) or 0 for x in xs) / len(xs)) if xs else 0,
        'min': lambda xs: min((x for x in xs if x is not None), default=None),
        'max': lambda xs: max((x for x in xs if x is not None), default=None),
        'first': lambda xs: xs[0] if xs else None,
    }

    @staticmethod
    def execute_task(task: ProcessingTask):
        """执行数据处理任务（自动分单/多 sheet 模式）"""
        task_start_time = time.monotonic()
        try:
            task.status = 'running'
            task.started_at = timezone.now()
            task.save(update_fields=['status', 'started_at'])

            mapping = task.mapping
            target_sheets = list(mapping.target_sheets.all().order_by('sort_order', 'id'))

            logger.info('任务开始执行', extra={
                'task_id': task.id,
                'mapping_id': mapping.id,
                'multi_sheet': bool(target_sheets),
                'sheet_count': len(target_sheets),
            })

            if target_sheets:
                DataProcessingService._execute_multi_sheet(task, mapping, target_sheets)
            else:
                DataProcessingService._execute_single_sheet(task, mapping)

            task.status = 'completed'
            task.completed_at = timezone.now()
            task.save(update_fields=[
                'status', 'completed_at', 'processed_rows',
                'success_rows', 'error_rows', 'result_file',
            ])

            task_duration = time.monotonic() - task_start_time
            logger.info('任务完成', extra={
                'task_id': task.id,
                'mapping_id': mapping.id,
                'total_rows': task.total_rows,
                'success_rows': task.success_rows,
                'error_rows': task.error_rows,
                'duration_seconds': round(task_duration, 3),
            })
            return True

        except Exception as e:
            task_duration = time.monotonic() - task_start_time
            logger.exception(f'任务执行失败: task_id={task.id}', extra={
                'task_id': task.id,
                'duration_seconds': round(task_duration, 3),
            })
            task.status = 'failed'
            task.error_message = str(e)
            task.completed_at = timezone.now()
            task.save(update_fields=['status', 'error_message', 'completed_at'])
            return False

    # ---------------- 单 sheet 模式（保留兼容） ----------------

    @staticmethod
    def _execute_single_sheet(task, mapping):
        """单 sheet 模式：沿用旧逻辑"""
        # 读取源文件数据
        source_data = ExcelService.get_sheet_data(
            mapping.source_file, mapping.source_sheet or None,
        )

        # 读取对照文件所有 Sheets 数据
        reference_sheets = {}
        if mapping.reference_file:
            reference_sheets = ExcelService.read_all_sheets(mapping.reference_file)

        # 获取字段映射，预构建对照表索引
        field_mappings = list(mapping.fields.all().order_by('sort_order'))
        ref_indexes = DataProcessingService._build_reference_indexes(
            field_mappings, reference_sheets,
        )

        task.total_rows = len(source_data['rows'])
        task.save(update_fields=['total_rows'])

        target_headers = [fm.target_field for fm in field_mappings]
        target_rows = []

        for row_idx, source_row in enumerate(source_data['rows']):
            try:
                result_rows = DataProcessingService._process_row(
                    source_row, source_data['headers'],
                    reference_sheets, field_mappings, ref_indexes,
                    upstream_sheet_outputs=None,
                )
                target_rows.extend(result_rows)
                task.success_rows += 1
            except Exception as e:
                task.error_rows += 1
                logger.debug(f'处理第{row_idx+1}行失败: {e}')

            task.processed_rows = row_idx + 1
            if row_idx % DataProcessingService.PROGRESS_SAVE_INTERVAL == 0:
                task.save(update_fields=['processed_rows', 'success_rows', 'error_rows'])

        result_file = DataProcessingService._create_result_file_single(
            target_headers, target_rows,
        )
        task.result_file.save(
            f"{sanitize_filename(task.name)}_{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx",
            result_file,
        )

    # ---------------- 多 sheet 模式 ----------------

    @staticmethod
    def _execute_multi_sheet(task, mapping, target_sheets):
        """多 sheet 模式：按血缘 DAG 拓扑逐 sheet 执行，最终合并成单一 xlsx"""
        # 血缘边
        edges = list(
            SheetLineage.objects.filter(mapping=mapping)
            .values_list('upstream_id', 'downstream_id')
        )
        ordered_sheets = topo_sort_sheets(target_sheets, edges)

        # Req 5.2: 环检测独立，向 task.error_message 追加警告，不阻塞执行
        if detect_cycle(target_sheets, edges):
            append_task_error(task, '检测到 Sheet 血缘环')

        # 读取源文件全部 sheet 数据（用于按 sheet 级映射选择源）
        all_source_sheets = {}
        if mapping.source_file:
            all_source_sheets = ExcelService.read_all_sheets(mapping.source_file)
        default_source_name = mapping.source_sheet or (next(iter(all_source_sheets), '') if all_source_sheets else '')

        # 对照文件
        reference_sheets = {}
        if mapping.reference_file:
            reference_sheets = ExcelService.read_all_sheets(mapping.reference_file)

        # 为每个目标 sheet 初始化 TaskSheetResult（Req 9.1）
        result_by_sheet_id = {}
        for order, ts in enumerate(ordered_sheets):
            tsr, _ = TaskSheetResult.objects.update_or_create(
                task=task, target_sheet=ts,
                defaults={
                    'sheet_name': ts.sheet_name,
                    'status': TaskSheetResult.Status.PENDING,
                    'execution_order': order,
                    'total_rows': 0,
                    'success_rows': 0,
                    'error_rows': 0,
                    'error_message': '',
                    'started_at': None,
                    'completed_at': None,
                    'duration_ms': 0,
                },
            )
            result_by_sheet_id[ts.id] = tsr

        # 每个 sheet 处理后的输出（headers + rows + 字段->列索引），供下游引用
        sheet_outputs = {}  # sheet_id -> {'sheet_name', 'headers', 'rows', 'col_index'}

        total_all = 0
        success_all = 0
        error_all = 0

        for ts in ordered_sheets:
            tsr = result_by_sheet_id[ts.id]

            # Req 2.5 / 9.6: disabled → SKIPPED 且跳过数据处理
            if ts.status == MappingTargetSheet.Status.DISABLED:
                tsr.status = TaskSheetResult.Status.SKIPPED
                tsr.completed_at = timezone.now()
                tsr.save(update_fields=['status', 'completed_at'])
                continue

            sheet_start = time.monotonic()
            tsr.status = TaskSheetResult.Status.RUNNING
            tsr.started_at = timezone.now()                                   # Req 9.3
            tsr.save(update_fields=['status', 'started_at'])

            try:
                # 决定该 sheet 使用的源数据
                src_name = ts.source_sheet or default_source_name             # Req 1.4
                source_data = all_source_sheets.get(src_name, {'headers': [], 'rows': []})

                field_mappings = list(ts.fields.all().order_by('sort_order'))
                ref_indexes = DataProcessingService._build_reference_indexes(
                    field_mappings, reference_sheets,
                )

                headers = [fm.target_field for fm in field_mappings]
                target_rows = []

                row_count = len(source_data['rows'])
                tsr.total_rows = row_count
                tsr.save(update_fields=['total_rows'])

                for row_idx, source_row in enumerate(source_data['rows']):
                    try:
                        result_rows = DataProcessingService._process_row(
                            source_row, source_data['headers'],
                            reference_sheets, field_mappings, ref_indexes,
                            upstream_sheet_outputs=sheet_outputs,             # Req 8.3
                        )
                        target_rows.extend(result_rows)
                        tsr.success_rows += 1
                    except Exception as e:
                        tsr.error_rows += 1
                        logger.debug(f'Sheet[{ts.sheet_name}] 第{row_idx+1}行失败: {e}')

                    if row_idx % DataProcessingService.PROGRESS_SAVE_INTERVAL == 0:
                        tsr.save(update_fields=['success_rows', 'error_rows'])  # Req 9.5

                # Req 6.7: 零行源数据 + 存在 default / computed / cross_sheet_ref 字段 → 生成 1 行
                if not source_data['rows'] and field_mappings and any(
                    fm.field_type in (MappingField.FieldType.DEFAULT,
                                      MappingField.FieldType.CROSS_SHEET_REF,
                                      MappingField.FieldType.COMPUTED)
                    for fm in field_mappings
                ):
                    try:
                        single_rows = DataProcessingService._process_row(
                            [], [], reference_sheets, field_mappings, ref_indexes,
                            upstream_sheet_outputs=sheet_outputs,
                        )
                        target_rows.extend(single_rows)
                        tsr.success_rows += 1
                        tsr.total_rows = 1
                    except Exception as e:
                        tsr.error_rows += 1
                        logger.debug(f'Sheet[{ts.sheet_name}] 单行聚合失败: {e}')

                sheet_outputs[ts.id] = {
                    'sheet_name': ts.sheet_name,
                    'headers': headers,
                    'rows': target_rows,
                    'col_index': {h: i for i, h in enumerate(headers)},
                }

                tsr.status = TaskSheetResult.Status.COMPLETED
            except Exception as e:                                            # Req 8.5
                logger.exception(f'Sheet 执行失败: {ts.sheet_name}')
                tsr.status = TaskSheetResult.Status.FAILED
                tsr.error_message = str(e)
            finally:
                tsr.completed_at = timezone.now()                             # Req 9.4
                tsr.duration_ms = int((time.monotonic() - sheet_start) * 1000)
                tsr.save()

            success_all += tsr.success_rows
            error_all += tsr.error_rows
            total_all += tsr.total_rows

        # Req 8.4: 汇总到 ProcessingTask
        task.total_rows = total_all
        task.processed_rows = total_all
        task.success_rows = success_all
        task.error_rows = error_all

        # 生成结果 xlsx（可能基于目标模板）
        result_file = DataProcessingService._create_result_file_multi(
            mapping, ordered_sheets, sheet_outputs,
        )
        # Req 10.6: 使用 sanitize_filename 保障文件名合法
        task.result_file.save(
            f"{sanitize_filename(task.name)}_{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx",
            result_file,
        )

    # ---------------- 通用：对照表索引构建 ----------------

    @staticmethod
    def _build_reference_indexes(field_mappings, reference_sheets):
        """为 lookup 类型字段预构建对照表索引，避免每行重复遍历"""
        indexes = {}

        for fm in field_mappings:
            if fm.field_type not in ['lookup', 'source_ref_target']:
                continue
            if not fm.reference_sheet or fm.reference_sheet not in reference_sheets:
                continue

            cache_key = (fm.reference_sheet, fm.reference_name_column, fm.reference_code_column)
            if cache_key in indexes:
                continue

            ref_data = reference_sheets[fm.reference_sheet]
            headers = ref_data.get('headers', [])
            rows = ref_data.get('rows', [])

            name_col_idx = -1
            code_col_idx = -1
            for idx, header in enumerate(headers):
                if header == fm.reference_name_column:
                    name_col_idx = idx
                if header == fm.reference_code_column:
                    code_col_idx = idx

            if name_col_idx < 0 or code_col_idx < 0:
                continue

            index = {}
            for row in rows:
                if name_col_idx < len(row) and row[name_col_idx] is not None:
                    key = str(row[name_col_idx]).strip()
                    if code_col_idx < len(row):
                        index.setdefault(key, []).append(row[code_col_idx])

            indexes[cache_key] = index

        return indexes

    # ---------------- 行级处理 ----------------

    @staticmethod
    def _process_row(source_row, source_headers, reference_sheets, field_mappings,
                     ref_indexes=None, upstream_sheet_outputs=None):
        """处理单行数据，可能返回多行结果。

        upstream_sheet_outputs 为 {sheet_id: {headers, rows, col_index}}，
        用于 cross_sheet_ref 和 computed 表达式中的 {sheet.field} 语法。
        """
        source_dict = {
            header: source_row[idx]
            for idx, header in enumerate(source_headers)
            if idx < len(source_row)
        }

        field_values = []
        field_is_list = []

        for fm in field_mappings:
            value = None
            is_list = False

            if fm.field_type in ['direct', 'source_to_target']:
                if fm.source_field and fm.source_field in source_dict:
                    value = source_dict[fm.source_field]
                elif 0 <= fm.source_field_index < len(source_row):
                    value = source_row[fm.source_field_index]

            elif fm.field_type in ['lookup', 'source_ref_target']:
                lookup_key = None
                if fm.source_field and fm.source_field in source_dict:
                    lookup_key = source_dict[fm.source_field]
                elif 0 <= fm.source_field_index < len(source_row):
                    lookup_key = source_row[fm.source_field_index]

                if lookup_key is not None and fm.reference_sheet:
                    cache_key = (fm.reference_sheet, fm.reference_name_column, fm.reference_code_column)

                    if ref_indexes and cache_key in ref_indexes:
                        matched = ref_indexes[cache_key].get(str(lookup_key).strip(), [])
                    elif fm.reference_sheet in reference_sheets:
                        matched = DataProcessingService._lookup_reference_v2(
                            lookup_key, reference_sheets[fm.reference_sheet],
                            fm.reference_name_column, fm.reference_code_column,
                        )
                    else:
                        matched = []

                    if len(matched) > 1:
                        value = matched
                        is_list = True
                    else:
                        value = matched[0] if matched else lookup_key
                else:
                    value = lookup_key

            elif fm.field_type == 'computed':
                if fm.compute_expression:
                    value = DataProcessingService._evaluate_expression(
                        fm.compute_expression, source_dict, upstream_sheet_outputs,
                    )

            elif fm.field_type == 'default':
                value = fm.default_value

            elif fm.field_type == MappingField.FieldType.CROSS_SHEET_REF:
                value = DataProcessingService._resolve_cross_sheet_ref(
                    fm, source_dict, upstream_sheet_outputs,
                )

            elif fm.field_type == 'ref_to_target':
                if fm.reference_sheet and fm.reference_sheet in reference_sheets:
                    ref_data = reference_sheets[fm.reference_sheet]
                    ref_idx = fm.reference_field_index
                    if ref_data['rows'] and 0 <= ref_idx < len(ref_data['rows'][0]):
                        value = ref_data['rows'][0][ref_idx]

            if fm.transform_rule:
                if is_list:
                    value = [DataProcessingService._apply_transform(v, fm.transform_rule) for v in value]
                else:
                    value = DataProcessingService._apply_transform(value, fm.transform_rule)

            field_values.append(value)
            field_is_list.append(is_list)

        if not any(field_is_list):
            return [field_values]

        expanded = [
            val if field_is_list[i] else [val]
            for i, val in enumerate(field_values)
        ]
        return [list(combo) for combo in product(*expanded)]

    # ---------------- 跨 sheet 引用 ----------------

    @staticmethod
    def _resolve_cross_sheet_ref(fm, source_dict, upstream_sheet_outputs):
        """解析 cross_sheet_ref 类型的字段值。

        读取 source_target_sheet + source_target_field + aggregation。
        """
        if not upstream_sheet_outputs or not fm.source_target_sheet_id:
            return None
        upstream = upstream_sheet_outputs.get(fm.source_target_sheet_id)
        if not upstream:
            return None
        col_idx = upstream['col_index'].get(fm.source_target_field, -1)
        if col_idx < 0:
            return None
        column_values = [row[col_idx] for row in upstream['rows'] if col_idx < len(row)]

        agg = (fm.aggregation or '').lower().strip()
        if agg and agg in DataProcessingService.AGG_FUNCS:
            return DataProcessingService.AGG_FUNCS[agg](column_values)

        # 无聚合：默认取第一行（适合 default/静态映射场景）
        return column_values[0] if column_values else None

    # ---------------- lookup helpers ----------------

    @staticmethod
    def _lookup_reference_v2(lookup_key, reference_data, name_column, code_column):
        """在对照表中查找所有匹配值"""
        if lookup_key is None:
            return [lookup_key]

        headers = reference_data.get('headers', [])
        rows = reference_data.get('rows', [])

        name_col_idx = -1
        code_col_idx = -1
        for idx, header in enumerate(headers):
            if header == name_column:
                name_col_idx = idx
            if header == code_column:
                code_col_idx = idx

        if name_col_idx < 0 or code_col_idx < 0:
            return [lookup_key]

        lookup_str = str(lookup_key).strip()
        matched_values = []
        for row in rows:
            if name_col_idx < len(row):
                cell_value = row[name_col_idx]
                if cell_value is not None and str(cell_value).strip() == lookup_str:
                    if code_col_idx < len(row):
                        matched_values.append(row[code_col_idx])

        return matched_values if matched_values else [lookup_key]

    @staticmethod
    def _apply_transform(value, rule):
        """应用数据转换规则"""
        if not rule or value is None:
            return value

        rule_type = rule.get('type')
        str_value = str(value)

        if rule_type == 'uppercase':
            return str_value.upper()
        elif rule_type == 'lowercase':
            return str_value.lower()
        elif rule_type == 'prefix':
            return f"{rule.get('value', '')}{value}"
        elif rule_type == 'suffix':
            return f"{value}{rule.get('value', '')}"
        elif rule_type == 'replace':
            return str_value.replace(rule.get('old', ''), rule.get('new', ''))
        return value

    @staticmethod
    def _evaluate_expression(expression, source_dict, upstream_sheet_outputs=None):
        """计算表达式，支持 {字段名} 和 {sheet.字段} 引用。

        Req 7.6 不变式：任何异常都必须被捕获并返回 None（不抛给上层 _process_row）。
        这里加一层顶层 try/except 作为最终兜底，即使内部 replace_field / safe_eval_expr
        在未来演进中意外抛错也能维持契约。
        """
        if not expression:
            return None
        try:
            return DataProcessingService._evaluate_expression_impl(
                expression, source_dict, upstream_sheet_outputs,
            )
        except Exception as e:  # Req 7.6
            logger.debug(f'表达式求值失败: {expression!r} -> {e}')
            return None

    @staticmethod
    def _evaluate_expression_impl(expression, source_dict, upstream_sheet_outputs=None):
        """实际求值逻辑，假设调用方已用 _evaluate_expression 包了一层 try/except。"""
        # 构建 sheet_name -> col_index 和 rows 的索引（用于 {sheet.field}）
        upstream_by_name = {}
        if upstream_sheet_outputs:
            for out in upstream_sheet_outputs.values():
                upstream_by_name[out['sheet_name']] = out

        def replace_field(match):
            token = match.group(1)
            value = None
            if '.' in token and upstream_by_name:
                sheet_name, field_name = token.split('.', 1)
                out = upstream_by_name.get(sheet_name.strip())
                if out:
                    col_idx = out['col_index'].get(field_name.strip(), -1)
                    if col_idx >= 0 and out['rows']:
                        # 默认取第一行的该字段（与 cross_sheet_ref 无聚合语义一致）
                        row0 = out['rows'][0]
                        if col_idx < len(row0):
                            value = row0[col_idx]
            else:
                value = source_dict.get(token)

            if value is None:
                return '0'
            if isinstance(value, (int, float)):
                return str(value)
            try:
                return str(float(value))
            except (ValueError, TypeError):
                return '0'

        result_expr = _FIELD_PATTERN.sub(replace_field, expression)

        result = safe_eval_expr(result_expr)
        if result is None:
            return None
        if isinstance(result, float):
            return int(result) if result == int(result) else round(result, 6)
        return result

    # ---------------- 结果文件生成 ----------------

    @staticmethod
    def _create_result_file_single(headers, rows):
        """创建单 sheet 结果 Excel 文件（向后兼容）"""
        wb = Workbook()
        ws = wb.active
        ws.title = '处理结果'
        ws.append(headers)
        for row in rows:
            ws.append(row)

        output = io.BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)
        return ContentFile(output.read())

    @staticmethod
    def _create_result_file_multi(mapping, ordered_sheets, sheet_outputs):
        """创建多 sheet 结果 Excel：优先基于目标模板，失败降级为新建工作簿（Req 10.5）。

        具体分支委托给 ``services_exporter`` 中的纯函数：
          - ``build_workbook_from_template``：模板加载成功时，按 Req 10.2 / 10.3 / 10.4 填充
          - ``build_workbook_fresh``：无模板或模板加载失败的兜底分支（Req 10.1 / 10.5）
        """
        wb = None
        try:
            if mapping.target_template:
                try:
                    file_bytes = ExcelService._read_file_to_bytes(mapping.target_template)
                    template_wb = load_workbook(file_bytes)
                    wb = build_workbook_from_template(template_wb, ordered_sheets, sheet_outputs)
                except Exception as e:
                    # Req 10.5: 降级为新建 workbook
                    logger.warning(f'加载目标模板失败，回退为新建 workbook: {e}')
                    wb = None

            if wb is None:
                wb = build_workbook_fresh(ordered_sheets, sheet_outputs)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return ContentFile(output.read())
        finally:
            if wb is not None:
                wb.close()


def _to_number(x):
    """尝试将任意值转为数值；失败返回 None"""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return x
    try:
        return float(x)
    except (ValueError, TypeError):
        return None
