"""
多 Sheet 结果导出纯函数

从 DataProcessingService._create_result_file_multi 中抽离的两种工作簿构建模式：

- build_workbook_fresh          : 无模板场景，按拓扑顺序新建工作表
- build_workbook_from_template  : 有模板场景，保留模板中未配置的工作表并填充已配置的工作表
- sanitize_filename             : 文件名合法化

这些函数是"纯函数"：不访问数据库、不读写文件系统，便于单元测试和属性测试。
文件 IO（加载模板、写出 xlsx 字节流、容错降级）由上层 service 负责。

对应 Requirement 10.1 / 10.2 / 10.3 / 10.4。
"""
from __future__ import annotations

import re
from typing import Any, Dict, Iterable, List, Mapping

from openpyxl import Workbook
from openpyxl.workbook.workbook import Workbook as WorkbookType


# 单 Sheet 模式下的兜底标题（Req 17.2）
_FALLBACK_SHEET_TITLE = '处理结果'

# 非法文件名字符：Windows + POSIX 的超集
_ILLEGAL_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|]+')


def build_workbook_fresh(ordered_sheets, sheet_outputs: Mapping[int, Dict[str, Any]]) -> WorkbookType:
    """新建工作簿并按拓扑顺序填入每个有输出的目标 Sheet（Req 10.1）。

    Args:
        ordered_sheets: 拓扑排序后的 MappingTargetSheet 列表（顺序即写入顺序）。
        sheet_outputs: {sheet_id: {'sheet_name', 'headers', 'rows', 'col_index'}}。
            未出现在该字典中的 sheet 视为无输出（SKIPPED / FAILED），跳过不写。

    Returns:
        openpyxl.Workbook：
          - 按 ordered_sheets 顺序逐个创建同名工作表
          - 无任何输出时回退创建一张 `处理结果` 空表，保证 xlsx 合法

    本函数不触发任何文件 IO。
    """
    wb = Workbook()
    # 删除 openpyxl 自动创建的默认工作表
    default = wb.active
    wb.remove(default)

    for ts in ordered_sheets:
        out = sheet_outputs.get(ts.id)
        if not out:
            continue
        headers = list(out.get('headers') or [])
        rows = out.get('rows') or []

        ws = wb.create_sheet(title=ts.sheet_name)
        ws.append(headers)
        for row in rows:
            ws.append(list(row))

    # 所有 sheet 均 SKIPPED / 无输出 → 留一张空表保证 workbook 合法
    if not wb.sheetnames:
        wb.create_sheet(title=_FALLBACK_SHEET_TITLE)

    return wb


def build_workbook_from_template(
    template_wb: WorkbookType,
    ordered_sheets,
    sheet_outputs: Mapping[int, Dict[str, Any]],
) -> WorkbookType:
    """基于已加载的模板工作簿，按拓扑顺序填入已配置的目标 Sheet。

    三种分支（与 Req 10.2 / 10.3 / 10.4 对齐）：
      1. 同名工作表 + 表头前 N 列匹配 → `delete_rows(2, ws.max_row - 1)` 后 `append(row)`
         → 保留原始表头样式 + MergedCell 区（Req 10.3）
      2. 同名工作表但表头不匹配 → 清空所有行后写入 headers + 数据（Req 10.4）
      3. 模板中不存在同名工作表 → 新建同名工作表并写入 headers + 数据
      4. 非同名（未配置）工作表保持不变（Req 10.2）

    Args:
        template_wb: 已通过 openpyxl.load_workbook 加载的模板 Workbook（会被就地修改）。
        ordered_sheets: 拓扑排序后的 MappingTargetSheet 列表。
        sheet_outputs: 同 build_workbook_fresh。

    Returns:
        原始 template_wb 对象（就地修改）。
    """
    for ts in ordered_sheets:
        out = sheet_outputs.get(ts.id)
        if not out:
            continue
        headers = list(out.get('headers') or [])
        rows = out.get('rows') or []

        if ts.sheet_name in template_wb.sheetnames:
            ws = template_wb[ts.sheet_name]
            first_row = [c.value for c in ws[1]] if ws.max_row >= 1 else []
            n = len(headers)

            if n > 0 and first_row[:n] == headers:
                # 表头匹配：仅清除数据区，保留原始表头样式（Req 10.3）
                if ws.max_row > 1:
                    ws.delete_rows(2, ws.max_row - 1)
                for row in rows:
                    ws.append(list(row))
            else:
                # 表头不匹配：全量清空后重写（Req 10.4）
                if ws.max_row > 0:
                    ws.delete_rows(1, ws.max_row)
                ws.append(headers)
                for row in rows:
                    ws.append(list(row))
        else:
            # 模板中无同名 sheet → 新建
            ws = template_wb.create_sheet(title=ts.sheet_name)
            ws.append(headers)
            for row in rows:
                ws.append(list(row))

    # 未在 ordered_sheets 中出现的工作表（Req 10.2）由于本函数从不触碰，自然保持不变
    return template_wb


def sanitize_filename(name: str) -> str:
    """将文件名中的非法字符（\\ / : * ? " < > |）替换为下划线。

    - 连续多个非法字符合并为单个下划线
    - 首尾空白被裁剪
    - 替换后为空串时回退为 'result'

    对应 Req 10.6 的稳健命名保障，供 `_create_result_file_multi` 与
    `_create_result_file_single` 在写 task.result_file 时调用。
    """
    if not name:
        return 'result'
    cleaned = _ILLEGAL_FILENAME_CHARS.sub('_', str(name)).strip()
    return cleaned or 'result'
